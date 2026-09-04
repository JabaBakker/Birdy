"""Financiële module (3.0): het register leeft in een Google Sheet in de Drive-hub
(`30 Financiën/Financieel overzicht`), documenten in de mappen eronder. Deze module leest
de Sheet (via Drive-export naar xlsx, dus zonder extra Google-rechten), rekent alles om naar
bedragen per maand, en levert de Geld-tab, de verrekening tussen de rekeningen en de
signalen (polis verloopt, rentevaste periode eindigt).

CLI voor het brein:
    python /app/agent/financien.py toon     compact overzicht van het register
    python /app/agent/financien.py maak     de Sheet (met kopregels en voorbeelden) aanmaken

Tabbladen en kolommen (kopregel = rij 1; hoofdletters/spaties maken niet uit):
  Rekeningen   : Naam · IBAN · Van · Doel                ← "Betaald van" verwijst naar deze namen
  Inkomsten    : Naam · Bedrag · Frequentie · Komt binnen op · Hoort bij · Herkenning · Notitie
  Vaste lasten : Naam · Categorie · Bedrag · Frequentie · Betaald van · Hoort bij · Betaaldag ·
                 Opzegtermijn · Einddatum · Herkenning · Document · Notitie
  Polissen     : Verzekering · Verzekeraar · Dekking · Premie · Frequentie · Eigen risico ·
                 Betaald van · Hoort bij · Einddatum · Opzegtermijn · Document · Notitie
  Hypotheek    : Deel · Verstrekker · Hoofdsom · Restschuld · Rente % · Rentevast tot ·
                 Aflossingsvorm · Maandlast · Waarvan rente · Waarvan aflossing · Einddatum ·
                 Document · Notitie
  Incidenteel  : Naam · Soort (inkomst/uitgave) · Bedrag · Wanneer (maand of datum) · Voorspelbaar (ja/nee) ·
                 Rekening · Hoort bij · Notitie   ← bonus, vakantiegeld, vakantie, aanslagen, reparaties
  Besparingen  : Voorstel · Besparing per maand · Categorie · Status (idee/onderzoeken/gepland/gedaan/afgewezen) ·
                 Bron · Datum · Notitie
  Variabele kosten: Naam · Budget · Frequentie · Betaald van · Hoort bij · Herkenning · Notitie
                 ← budgetten voor wat elke maand wisselt (boodschappen, uitjes, kleding, auto)
  Geldstromen  : Naam · Richting (in/uit) · Bedrag · Frequentie · Van · Naar · Hoort bij ·
                 Categorie · Uitleg      ← voor constructies zoals een lening bij familie
  Uitleg       : Onderwerp · Tekst · Bijgewerkt   ← door Birdy geschreven uitleg in gewone taal
Frequentie: maand · kwartaal · halfjaar · jaar · week · 2 weken · eenmalig.
Betaald van / Komt binnen op: een rekening uit het tabblad Rekeningen (of een persoonsnaam);
Hoort bij: een persoon of "gezamenlijk" (de pot). Verrekenen gebeurt tegenover de pot: wie privé
iets voor de pot betaalt krijgt dat terug; wie privé iets van de pot ontvangt, is dat de pot schuldig.
"Herkenning" = tegenrekening (IBAN) of trefwoord, voor het automatisch matchen van bankexports (3.1).
"Verrekend via" = "inleg" als een post al in de inleg-afspraak zit (bijv. de toeslag die privé binnenkomt,
of energie die privé betaald wordt): dan telt hij niet mee in de verrekening maar staat hij bij de constructie.
"""
from __future__ import annotations

import io
import logging
import os
import re
import sys
from datetime import date, datetime

log = logging.getLogger("fien.financien")

SHEET_MAP = "30 Financiën"
SHEET_NAAM = "Financieel overzicht"
SHEET_PAD = f"{SHEET_MAP}/{SHEET_NAAM}"
SUBMAPPEN = ("Hypotheek", "Verzekeringen", "Abonnementen & vaste lasten", "Belasting", "Bankexports")
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
SHEET_MIME = "application/vnd.google-apps.spreadsheet"

TABS: dict[str, list[str]] = {
    "Rekeningen": ["Naam", "IBAN", "Van", "Doel"],
    "Inkomsten": ["Naam", "Bedrag", "Frequentie", "Komt binnen op", "Hoort bij", "Herkenning", "Uitleg",
                  "Verrekend via", "Notitie"],
    "Vaste lasten": ["Naam", "Categorie", "Bedrag", "Frequentie", "Betaald van", "Hoort bij",
                     "Betaaldag", "Opzegtermijn", "Startdatum", "Einddatum", "Overstapdatum", "Herkenning",
                     "Verrekend via", "Document", "Notitie"],
    "Polissen": ["Verzekering", "Verzekeraar", "Dekking", "Premie", "Frequentie", "Eigen risico",
                 "Betaald van", "Hoort bij", "Startdatum", "Einddatum", "Overstapdatum", "Opzegtermijn",
                 "Herkenning", "Verrekend via", "Document", "Notitie"],
    "Incidenteel": ["Naam", "Soort", "Bedrag", "Wanneer", "Voorspelbaar", "Rekening", "Hoort bij", "Notitie"],
    "Besparingen": ["Voorstel", "Besparing per maand", "Categorie", "Status", "Bron", "Datum", "Notitie"],
    "Hypotheek": ["Deel", "Verstrekker", "Hoofdsom", "Restschuld", "Rente %", "Rentevast tot",
                  "Aflossingsvorm", "Maandlast", "Waarvan rente", "Waarvan aflossing", "Einddatum",
                  "Betaald van", "Document", "Notitie"],
    "Variabele kosten": ["Naam", "Budget", "Frequentie", "Betaald van", "Hoort bij", "Herkenning", "Notitie"],
    "Geldstromen": ["Naam", "Richting", "Bedrag", "Frequentie", "Van", "Naar", "Hoort bij",
                    "Categorie", "Uitleg"],
    "Uitleg": ["Onderwerp", "Tekst", "Bijgewerkt"],
}

VOORBEELDEN: dict[str, list[list]] = {
    "Rekeningen": [["Gezamenlijk (voorbeeld)", "NL00BANK0000000000", "gezamenlijk", "vaste lasten gezin"]],
    "Inkomsten": [["Salaris (voorbeeld)", 3000, "maand", "Privé A", "A", "", ""]],
    "Vaste lasten": [
        ["Energie (voorbeeld)", "Wonen", 180, "maand", "gezamenlijk", "gezamenlijk", 1, "1 maand", "", "", "", "vervang of verwijder deze voorbeeldregel"],
        ["Netflix (voorbeeld)", "Abonnementen", 15.99, "maand", "Privé A", "gezamenlijk", 12, "1 maand", "", "", "", ""],
    ],
    "Polissen": [
        ["Autoverzekering (voorbeeld)", "—", "WA + casco", 62, "maand", 150, "Privé A", "gezamenlijk", "31-12-2026", "1 maand", "", "", ""],
    ],
    "Hypotheek": [
        ["Deel 1 (voorbeeld)", "—", 300000, 250000, 3.1, "01-07-2030", "annuïteit", 1450, 640, 810, "01-07-2050", "", ""],
    ],
    "Variabele kosten": [["Boodschappen (voorbeeld)", 600, "maand", "gezamenlijk", "gezamenlijk", "", "schatting"]],
    "Geldstromen": [
        ["Lening vader", "uit", 600, "maand", "gezamenlijk", "vader", "gezamenlijk", "Familie", "Maandelijkse betaling aan vader voor de lening."],
        ["Lening vader", "in", 1800, "kwartaal", "vader", "gezamenlijk", "gezamenlijk", "Familie", "Elk kwartaal komt er 1800 terug; netto per maand dus 0."],
    ],
    "Uitleg": [],
}

# uitleg in gewone taal van regelingen die als inkomst binnenkomen; (trefwoorden in de naam,
# titel, tekst). {bedrag}, {rekening} en {freq} worden ingevuld met de eigen cijfers.
REGELINGEN: list[tuple[tuple[str, ...], str, str]] = [
    (("kinderopvangtoeslag", "opvangtoeslag"), "Kinderopvangtoeslag",
     "Een bijdrage van de overheid in de kosten van kinderopvang. De hoogte hangt af van jullie gezamenlijke "
     "inkomen, het aantal opvanguren en het uurtarief (tot een maximum). Je krijgt hem maandelijks vooruit als "
     "voorschot; na afloop van het jaar rekent de Belastingdienst definitief af, dus bij een hoger inkomen of "
     "minder uren kan een deel terug moeten. Bij jullie: {bedrag} per {freq} op {rekening}, terwijl de opvang "
     "zelf van de pot gaat — daarom telt dit mee in de verrekening."),
    (("kinderbijslag",), "Kinderbijslag",
     "Een vast bedrag per kind van de Sociale Verzekeringsbank, los van jullie inkomen. Het wordt per kwartaal "
     "achteraf betaald en stijgt als een kind 6 en 12 wordt. Bij jullie: {bedrag} per {freq} op {rekening}."),
    (("kindgebonden",), "Kindgebonden budget",
     "Een inkomensafhankelijke toeslag per kind van de Belastingdienst, maandelijks vooruit als voorschot en "
     "achteraf definitief afgerekend. Bij jullie: {bedrag} per {freq} op {rekening}."),
    (("schenking",), "Schenking",
     "Een gift die tot een jaarlijks vrijgesteld bedrag belastingvrij is (ouders → kind kent een hogere "
     "vrijstelling). Bij jullie: {bedrag} per {freq} op {rekening}."),
    (("zorgtoeslag",), "Zorgtoeslag",
     "Bijdrage in de zorgpremie voor lagere inkomens, maandelijks vooruit, achteraf afgerekend. Bij jullie: "
     "{bedrag} per {freq} op {rekening}."),
    (("huurtoeslag",), "Huurtoeslag",
     "Bijdrage in de huur voor lagere inkomens, maandelijks vooruit, achteraf afgerekend. Bij jullie: {bedrag} "
     "per {freq} op {rekening}."),
]


def regeling_uitleg(naam: str, bedrag: float, freq: str, rekening: str) -> tuple[str, str]:
    n = naam.lower()
    for woorden, titel, tekst in REGELINGEN:
        if any(w in n for w in woorden):
            return titel, tekst.format(bedrag=f"€ {eur(bedrag)}", freq=freq or "maand", rekening=rekening or "?")
    return "", ""


# frequentie → factor naar 'per maand'
FREQ = {"maand": 1.0, "maandelijks": 1.0, "per maand": 1.0, "kwartaal": 1 / 3, "per kwartaal": 1 / 3,
        "2 maanden": 1 / 2, "twee maanden": 1 / 2, "4 maanden": 1 / 4, "vier maanden": 1 / 4,
        "halfjaar": 1 / 6, "half jaar": 1 / 6, "jaar": 1 / 12, "jaarlijks": 1 / 12, "per jaar": 1 / 12,
        "week": 52 / 12, "wekelijks": 52 / 12, "2 weken": 26 / 12, "twee weken": 26 / 12,
        "eenmalig": 0.0, "": 1.0}

WOORDENLIJST: list[tuple[str, str]] = [
    ("Rente", "Wat je de bank betaalt om geld te mogen lenen: een percentage van de schuld per jaar. Bij 3% op 250.000 euro is dat ongeveer 625 euro per maand."),
    ("Aflossing", "Het deel van je maandbedrag waarmee je de schuld zelf kleiner maakt. Elke euro aflossing is een euro minder schuld, en daardoor volgend jaar iets minder rente."),
    ("Annuïteit", "Elke maand hetzelfde bedrag. In het begin is dat vooral rente en weinig aflossing; later draait dat om. Fijn voorspelbaar."),
    ("Lineair", "Elke maand hetzelfde stuk aflossing, dus het totale bedrag begint hoger en wordt elk jaar lager, omdat je steeds minder rente betaalt."),
    ("Rentevaste periode", "Hoe lang je rente vaststaat. Loopt die af, dan krijg je een nieuw percentage (hoger of lager). Een half jaar van tevoren beginnen met vergelijken loont."),
    ("Restschuld", "Wat je op dit moment nog aan de bank moet. De maandlast zegt weinig; dit getal is de echte stand."),
    ("NHG", "Nationale Hypotheek Garantie: een vangnet als je de hypotheek niet meer kunt betalen. Levert meestal ook een iets lagere rente op."),
    ("Eigen risico", "Het bedrag dat je bij schade eerst zelf betaalt voordat de verzekering betaalt. Hoger eigen risico is een lagere premie, maar meer pijn bij schade."),
    ("Premie", "Wat je betaalt voor een verzekering, per maand of per jaar."),
    ("Opzegtermijn", "Hoe lang van tevoren je moet opzeggen. Mis je die, dan zit je er vaak weer een jaar aan vast."),
    ("Vaste lasten", "Alles wat elke maand sowieso weggaat: wonen, energie, verzekeringen, abonnementen. Wat overblijft is vrij te besteden."),
    ("Verrekenen", "Als iets van de ene rekening is betaald maar van de gezamenlijke of van de ander had moeten komen, moet dat rechtgetrokken worden. Het dashboard rekent uit wie wat aan wie moet."),
]


# ── helpers ──────────────────────────────────────────────────────────────────

def eur(n: float) -> str:
    """Nederlandse notatie zonder decimalen: 1.142 · 78"""
    return f"{n:,.0f}".replace(",", ".")


def _bedrag(v) -> float:
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace("€", "").replace(" ", "").replace(".", "").replace(",", ".") \
        if re.search(r"\d\.\d{3}(,|$)", str(v)) else str(v).replace("€", "").replace(" ", "").replace(",", ".")
    try:
        return float(re.sub(r"[^0-9.\-]", "", s) or 0)
    except ValueError:
        return 0.0


def _datum(v) -> date | None:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v or "").strip()
    for fmt in ("%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y", "%m-%Y", "%d-%m-%y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


MAANDEN = {"jan": 1, "feb": 2, "mrt": 3, "maa": 3, "apr": 4, "mei": 5, "jun": 6, "jul": 7, "aug": 8,
           "sep": 9, "okt": 10, "nov": 11, "dec": 12}


def _datum_dagen_maand(tekst: str, vandaag: date) -> int | None:
    """'mei' / 'juli 2027' / '15-05-2026' → dagen tot dat moment (maandnaam = de 1e van die maand,
    eerstvolgende keer)."""
    d = _datum(tekst)
    if d:
        return (d - vandaag).days
    t = str(tekst or "").strip().lower()[:3]
    if t in MAANDEN:
        m = MAANDEN[t]
        jaar = vandaag.year if m >= vandaag.month else vandaag.year + 1
        return (date(jaar, m, 1) - vandaag).days
    return None


def _termijn_dagen(v) -> int:
    """'1 maand' → 30, '3 maanden' → 90, '14 dagen' → 14, '' → 30 (veilige aanname)."""
    s = str(v or "").lower()
    m = re.search(r"(\d+)", s)
    n = int(m.group(1)) if m else 1
    if "dag" in s:
        return n
    if "week" in s:
        return n * 7
    if "jaar" in s:
        return n * 365
    return n * 30


def per_maand(bedrag, frequentie) -> float:
    f = FREQ.get(str(frequentie or "").strip().lower())
    if f is None:
        f = 1.0
    return round(_bedrag(bedrag) * f, 2)


def _norm(s) -> str:
    return re.sub(r"[^a-z0-9%]", "", str(s or "").lower())


def _rijen(ws, kolommen: list[str]) -> list[dict]:
    """Werkblad → lijst dicts met de opgegeven kolomnamen (kopregel wordt herkend, volgorde vrij)."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    kop = [_norm(c) for c in rows[0]]
    idx = {}
    for naam in kolommen:
        n = _norm(naam)
        for i, k in enumerate(kop):
            if k == n or (n and k.startswith(n)):
                idx[naam] = i
                break
    out = []
    for r in rows[1:]:
        if not r or all(c in (None, "") for c in r):
            continue
        item = {naam: (r[i] if i < len(r) else None) for naam, i in idx.items()}
        eerste = item.get(kolommen[0])
        if eerste in (None, "") or str(eerste).lower().startswith("(voorbeeld"):
            continue
        out.append(item)
    return out


def _wie(v) -> str:
    s = str(v or "").strip()
    return "gezamenlijk" if not s or s.lower() in ("gezamenlijk", "samen", "gedeeld", "beiden") else s


# ── Sheet aanmaken / lezen ───────────────────────────────────────────────────

def _werkboek(voorbeelden: bool = True):
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    wb.remove(wb.active)
    for tab, kolommen in TABS.items():
        ws = wb.create_sheet(tab)
        ws.append(kolommen)
        for c in ws[1]:
            c.font = Font(bold=True)
        if voorbeelden:
            for rij in VOORBEELDEN.get(tab, []):
                ws.append(rij)
        ws.freeze_panes = "A2"
        for kol in ws.columns:
            ws.column_dimensions[kol[0].column_letter].width = 18
    return wb


def maak_sheet() -> str:
    """Mappen + Sheet in Drive aanmaken (bestaat de Sheet al, dan blijft hij staan). Geeft de link."""
    from googleapiclient.http import MediaIoBaseUpload

    from . import gdrive

    svc = gdrive._service()
    for sub in SUBMAPPEN:
        gdrive._ensure_folder(svc, f"{SHEET_MAP}/{sub}")
    bestaand = gdrive._resolve(svc, SHEET_PAD, must_exist=False)
    if bestaand:
        return bestaand.get("webViewLink", "") or f"bestaat al: {bestaand['id']}"
    buf = io.BytesIO()
    _werkboek().save(buf)
    buf.seek(0)
    made = svc.files().create(
        body={"name": SHEET_NAAM, "mimeType": SHEET_MIME, "parents": [gdrive._ensure_folder(svc, SHEET_MAP)]},
        media_body=MediaIoBaseUpload(buf, mimetype=XLSX_MIME, resumable=False),
        fields="id, webViewLink",
    ).execute()
    return made.get("webViewLink", "")


def _laad_werkboek():
    from openpyxl import load_workbook

    from . import gdrive

    svc = gdrive._service()
    node = gdrive._resolve(svc, SHEET_PAD, must_exist=False)
    if not node:
        return None, ""
    data = svc.files().export(fileId=node["id"], mimeType=XLSX_MIME).execute()
    link = svc.files().get(fileId=node["id"], fields="webViewLink").execute().get("webViewLink", "")
    return load_workbook(io.BytesIO(data), data_only=True), link


def register(wb=None, vandaag: date | None = None, verdeling: dict | None = None) -> dict:
    """Alles omgerekend naar per maand + verrekening (pot-model) + signalen. `wb` = openpyxl-
    werkboek (voor tests); anders wordt de Sheet uit Drive geladen. `verdeling` is niet meer
    in gebruik (compat)."""
    vandaag = vandaag or date.today()
    link = ""
    if wb is None:
        wb, link = _laad_werkboek()
        if wb is None:
            return {"beschikbaar": False, "link": "", "woordenlijst": WOORDENLIJST}

    def ws(naam):
        for n in wb.sheetnames:
            if _norm(n) == _norm(naam):
                return wb[n]
        return None

    def lees(tab):
        w = ws(tab)
        return _rijen(w, TABS[tab]) if w is not None else []

    # rekeningen: naam → van wie (zodat "betaald van ABN Jaap" telt als Jaap)
    rekeningen = []
    rek_van: dict[str, str] = {}
    for r in lees("Rekeningen"):
        naam, van = str(r.get("Naam")).strip(), _wie(r.get("Van"))
        rekeningen.append({"naam": naam, "iban": str(r.get("IBAN") or ""), "van": van, "doel": str(r.get("Doel") or "")})
        rek_van[naam.lower()] = van

    def eigenaar(v) -> str:
        """rekeningnaam of persoonsnaam → persoon/gezamenlijk"""
        w = _wie(v)
        return rek_van.get(w.lower(), w)

    # inkomsten
    inkomsten = []
    for r in lees("Inkomsten"):
        inkomsten.append({
            "naam": str(r.get("Naam")), "bedrag": _bedrag(r.get("Bedrag")), "frequentie": str(r.get("Frequentie") or "maand"),
            "per_maand": per_maand(r.get("Bedrag"), r.get("Frequentie")),
            "komt_binnen_op": _wie(r.get("Komt binnen op")), "ontvanger": eigenaar(r.get("Komt binnen op")),
            "hoort_bij": _wie(r.get("Hoort bij")), "notitie": str(r.get("Notitie") or ""),
            "herkenning": str(r.get("Herkenning") or ""), "verrekend_via": str(r.get("Verrekend via") or "").strip().lower(),
        })
        titel, tekst = regeling_uitleg(inkomsten[-1]["naam"], inkomsten[-1]["bedrag"], inkomsten[-1]["frequentie"],
                                       inkomsten[-1]["komt_binnen_op"])
        eigen = str(r.get("Uitleg") or "").strip()
        inkomsten[-1]["regeling"] = titel
        inkomsten[-1]["uitleg"] = eigen or tekst

    # vaste lasten
    lasten = []
    for r in lees("Vaste lasten"):
        pm = per_maand(r.get("Bedrag"), r.get("Frequentie"))
        eind = _datum(r.get("Einddatum"))
        lasten.append({
            "naam": str(r.get("Naam")), "categorie": str(r.get("Categorie") or "Overig"),
            "bedrag": _bedrag(r.get("Bedrag")), "frequentie": str(r.get("Frequentie") or "maand"),
            "per_maand": pm, "betaald_van": _wie(r.get("Betaald van")), "betaler": eigenaar(r.get("Betaald van")),
            "hoort_bij": _wie(r.get("Hoort bij")),
            "betaaldag": r.get("Betaaldag") or "", "opzegtermijn": str(r.get("Opzegtermijn") or ""),
            "einddatum": eind.isoformat() if eind else "", "dagen": (eind - vandaag).days if eind else None,
            "herkenning": str(r.get("Herkenning") or ""), "verrekend_via": str(r.get("Verrekend via") or "").strip().lower(),
            "startdatum": _datum(r.get("Startdatum")).isoformat() if _datum(r.get("Startdatum")) else "",
            "overstapdatum": _datum(r.get("Overstapdatum")).isoformat() if _datum(r.get("Overstapdatum")) else "",
            "overstap_dagen": (_datum(r.get("Overstapdatum")) - vandaag).days if _datum(r.get("Overstapdatum")) else None,
            "document": str(r.get("Document") or ""), "notitie": str(r.get("Notitie") or ""),
        })
    variabel = []
    for r in lees("Variabele kosten"):
        variabel.append({
            "naam": str(r.get("Naam")), "bedrag": _bedrag(r.get("Budget")), "frequentie": str(r.get("Frequentie") or "maand"),
            "per_maand": per_maand(r.get("Budget"), r.get("Frequentie")),
            "betaald_van": _wie(r.get("Betaald van")), "betaler": eigenaar(r.get("Betaald van")),
            "hoort_bij": _wie(r.get("Hoort bij")), "herkenning": str(r.get("Herkenning") or ""),
            "notitie": str(r.get("Notitie") or ""),
        })
    incidenteel = []
    for r in lees("Incidenteel"):
        soort = "inkomst" if str(r.get("Soort") or "").lower().startswith("in") else "uitgave"
        voorsp = str(r.get("Voorspelbaar") or "ja").strip().lower() in ("ja", "j", "yes", "x", "true", "1")
        wanneer = str(r.get("Wanneer") or "")
        incidenteel.append({
            "naam": str(r.get("Naam")), "soort": soort, "bedrag": _bedrag(r.get("Bedrag")), "wanneer": wanneer,
            "dagen": _datum_dagen_maand(wanneer, vandaag), "voorspelbaar": voorsp,
            "rekening": _wie(r.get("Rekening")), "hoort_bij": _wie(r.get("Hoort bij")), "notitie": str(r.get("Notitie") or ""),
        })
    inc_in = round(sum(i["bedrag"] for i in incidenteel if i["soort"] == "inkomst" and i["voorspelbaar"]), 2)
    inc_uit = round(sum(i["bedrag"] for i in incidenteel if i["soort"] == "uitgave" and i["voorspelbaar"]), 2)
    besparingen = []
    for r in lees("Besparingen"):
        besparingen.append({
            "voorstel": str(r.get("Voorstel")), "per_maand": _bedrag(r.get("Besparing per maand")),
            "categorie": str(r.get("Categorie") or ""), "status": str(r.get("Status") or "idee").strip().lower() or "idee",
            "bron": str(r.get("Bron") or ""), "datum": str(r.get("Datum") or ""), "notitie": str(r.get("Notitie") or ""),
            "uit": "sheet",
        })
    per_cat: dict[str, float] = {}
    for l in lasten:
        per_cat[l["categorie"]] = round(per_cat.get(l["categorie"], 0) + l["per_maand"], 2)

    # polissen
    polissen = []
    for r in lees("Polissen"):
        eind = _datum(r.get("Einddatum"))
        polissen.append({
            "naam": str(r.get("Verzekering")), "verzekeraar": str(r.get("Verzekeraar") or ""),
            "dekking": str(r.get("Dekking") or ""), "premie": _bedrag(r.get("Premie")),
            "frequentie": str(r.get("Frequentie") or "maand"),
            "per_maand": per_maand(r.get("Premie"), r.get("Frequentie")),
            "eigen_risico": _bedrag(r.get("Eigen risico")), "betaald_van": _wie(r.get("Betaald van")),
            "betaler": eigenaar(r.get("Betaald van")),
            "hoort_bij": _wie(r.get("Hoort bij")), "einddatum": eind.isoformat() if eind else "",
            "dagen": (eind - vandaag).days if eind else None, "opzegtermijn": str(r.get("Opzegtermijn") or ""),
            "herkenning": str(r.get("Herkenning") or ""), "verrekend_via": str(r.get("Verrekend via") or "").strip().lower(),
            "startdatum": _datum(r.get("Startdatum")).isoformat() if _datum(r.get("Startdatum")) else "",
            "overstapdatum": _datum(r.get("Overstapdatum")).isoformat() if _datum(r.get("Overstapdatum")) else "",
            "overstap_dagen": (_datum(r.get("Overstapdatum")) - vandaag).days if _datum(r.get("Overstapdatum")) else None,
            "document": str(r.get("Document") or ""), "notitie": str(r.get("Notitie") or ""),
        })

    # hypotheek
    delen = []
    for r in lees("Hypotheek"):
        rv = _datum(r.get("Rentevast tot"))
        eind = _datum(r.get("Einddatum"))
        maandlast = _bedrag(r.get("Maandlast"))
        rente = _bedrag(r.get("Waarvan rente"))
        aflossing = _bedrag(r.get("Waarvan aflossing"))
        rest = _bedrag(r.get("Restschuld"))
        pct = _bedrag(r.get("Rente %"))
        if not rente and rest and pct:
            rente = round(rest * pct / 100 / 12, 2)  # schatting als de kolom leeg is
        if not aflossing and maandlast and rente:
            aflossing = round(maandlast - rente, 2)
        delen.append({
            "deel": str(r.get("Deel")), "verstrekker": str(r.get("Verstrekker") or ""),
            "hoofdsom": _bedrag(r.get("Hoofdsom")), "restschuld": rest, "rente_pct": pct,
            "rentevast_tot": rv.isoformat() if rv else "", "rentevast_dagen": (rv - vandaag).days if rv else None,
            "vorm": str(r.get("Aflossingsvorm") or ""), "maandlast": maandlast,
            "rente": rente, "aflossing": aflossing,
            "betaald_van": _wie(r.get("Betaald van")) if r.get("Betaald van") else "",
            "bekend": bool(rest or pct or rente),  # zonder restschuld/rente is er niets te verdelen
            "einddatum": eind.isoformat() if eind else "", "document": str(r.get("Document") or ""),
            "notitie": str(r.get("Notitie") or ""),
        })
    hyp = {
        "delen": delen,
        "maandlast": round(sum(d["maandlast"] for d in delen), 2),
        "rente": round(sum(d["rente"] for d in delen), 2),
        "aflossing": round(sum(d["aflossing"] for d in delen), 2),
        "restschuld": round(sum(d["restschuld"] for d in delen), 2),
        "hoofdsom": round(sum(d["hoofdsom"] for d in delen), 2),
        "bekend": any(d["bekend"] for d in delen),
    }

    # geldstromen → constructies (zelfde naam = één constructie: in − uit)
    stromen = []
    for r in lees("Geldstromen"):
        richting = "in" if str(r.get("Richting") or "").strip().lower().startswith("in") else "uit"
        stromen.append({
            "naam": str(r.get("Naam")), "richting": richting, "bedrag": _bedrag(r.get("Bedrag")),
            "frequentie": str(r.get("Frequentie") or "maand"),
            "per_maand": per_maand(r.get("Bedrag"), r.get("Frequentie")),
            "van": _wie(r.get("Van")), "betaler": eigenaar(r.get("Van")), "naar": _wie(r.get("Naar")),
            "hoort_bij": _wie(r.get("Hoort bij")),
            "categorie": str(r.get("Categorie") or ""), "uitleg": str(r.get("Uitleg") or ""),
        })
    constructies: dict[str, dict] = {}
    for s in stromen:
        c = constructies.setdefault(s["naam"], {"naam": s["naam"], "in_pm": 0.0, "uit_pm": 0.0, "stromen": [],
                                                "uitleg": "", "hoort_bij": s["hoort_bij"]})
        c["in_pm" if s["richting"] == "in" else "uit_pm"] += s["per_maand"]
        c["stromen"].append(s)
        if s["uitleg"] and s["uitleg"] not in c["uitleg"]:
            c["uitleg"] = (c["uitleg"] + " " + s["uitleg"]).strip()
    for c in constructies.values():
        c["in_pm"], c["uit_pm"] = round(c["in_pm"], 2), round(c["uit_pm"], 2)
        c["netto_pm"] = round(c["in_pm"] - c["uit_pm"], 2)

    # verrekening (pot-model): de gezamenlijke rekening is de pot. Wie privé iets betaalt dat
    # bij de pot hoort, krijgt dat van de pot terug (+). Wie privé iets ontvangt of laat betalen
    # dat van de pot is, is dat de pot schuldig (−). Inleg-afspraken tellen niet mee.
    personen = sorted({x for l in lasten + polissen for x in (l["betaler"], l["hoort_bij"]) if x != "gezamenlijk"}
                      | {i["hoort_bij"] for i in inkomsten if i["hoort_bij"] != "gezamenlijk"}
                      | {i["ontvanger"] for i in inkomsten if i["ontvanger"] != "gezamenlijk"}
                      | {s["hoort_bij"] for s in stromen if s["hoort_bij"] != "gezamenlijk"})
    saldo: dict[str, float] = {p: 0.0 for p in personen}  # + = pot is deze persoon geld schuldig
    regels = []

    def boek(wat, bedrag_pm, betaler, hoort_bij, soort="uitgave"):
        if bedrag_pm <= 0 or betaler == hoort_bij:
            return
        if soort == "inkomst":
            if hoort_bij == "gezamenlijk" and betaler in saldo:      # privé ontvangen, is van de pot
                saldo[betaler] -= bedrag_pm
                regels.append({"wat": wat, "bedrag": bedrag_pm, "wie": betaler, "richting": "aan_pot",
                               "tekst": f"komt binnen bij {betaler}, maar is van de pot"})
            elif betaler == "gezamenlijk" and hoort_bij in saldo:   # op de pot ontvangen, is van een persoon
                saldo[hoort_bij] += bedrag_pm
                regels.append({"wat": wat, "bedrag": bedrag_pm, "wie": hoort_bij, "richting": "van_pot",
                               "tekst": f"komt binnen op de pot, maar is van {hoort_bij}"})
            return
        if betaler == "gezamenlijk" and hoort_bij in saldo:          # pot betaalt iets van een persoon
            saldo[hoort_bij] -= bedrag_pm
            regels.append({"wat": wat, "bedrag": bedrag_pm, "wie": hoort_bij, "richting": "aan_pot",
                           "tekst": f"van {hoort_bij}, maar betaald door de pot"})
        elif hoort_bij == "gezamenlijk" and betaler in saldo:        # persoon betaalt iets van de pot
            saldo[betaler] += bedrag_pm
            regels.append({"wat": wat, "bedrag": bedrag_pm, "wie": betaler, "richting": "van_pot",
                           "tekst": f"van de pot, maar betaald door {betaler}"})
        elif betaler in saldo and hoort_bij in saldo:                # persoon betaalt iets van de ander
            saldo[betaler] += bedrag_pm
            saldo[hoort_bij] -= bedrag_pm
            regels.append({"wat": wat, "bedrag": bedrag_pm, "wie": betaler, "richting": "van_pot",
                           "tekst": f"van {hoort_bij}, maar betaald door {betaler}"})

    via_inleg = []  # posten die al in de inleg-afspraak verwerkt zitten (Verrekend via = inleg)
    for l in lasten:
        if l["verrekend_via"]:
            via_inleg.append({"wat": l["naam"], "bedrag": l["per_maand"], "richting": "uit", "wie": l["betaler"],
                              "tekst": f"{l['betaler']} betaalt dit privé voor de pot; zit in de inleg"})
            continue
        boek(l["naam"], l["per_maand"], l["betaler"], l["hoort_bij"])
    for p in polissen:
        if p["verrekend_via"]:
            via_inleg.append({"wat": p["naam"], "bedrag": p["per_maand"], "richting": "uit", "wie": p["betaler"],
                              "tekst": f"{p['betaler']} betaalt dit privé voor de pot; zit in de inleg"})
            continue
        boek(p["naam"], p["per_maand"], p["betaler"], p["hoort_bij"])
    for i in inkomsten:
        if i["verrekend_via"]:
            via_inleg.append({"wat": i["naam"], "bedrag": i["per_maand"], "richting": "in", "wie": i["ontvanger"],
                              "tekst": f"komt binnen bij {i['ontvanger']} maar is van de pot; zit in de inleg"})
            continue
        boek(i["naam"], i["per_maand"], i["ontvanger"], i["hoort_bij"], "inkomst")
    for st in stromen:
        if st["richting"] == "uit" and st["categorie"].lower() != "inleg":
            boek(st["naam"], st["per_maand"], st["betaler"], st["hoort_bij"])
        elif st["richting"] == "in" and st["categorie"].lower() != "inleg":
            boek(st["naam"], st["per_maand"], eigenaar(st["naar"]), st["hoort_bij"], "inkomst")
    saldo = {p: round(v, 2) for p, v in saldo.items()}
    delen_tekst = []
    for p, v in saldo.items():
        if v > 0.5:
            delen_tekst.append(f"de pot is {p} € {eur(v)} per maand schuldig")
        elif v < -0.5:
            delen_tekst.append(f"{p} is de pot € {eur(-v)} per maand schuldig")
    verreken_tekst = "; ".join(delen_tekst) if delen_tekst else "structureel in balans"
    verdeling = {}

    # uitleg-cache (door Birdy geschreven)
    uitleg = {str(r.get("Onderwerp")).strip().lower(): {"tekst": str(r.get("Tekst") or ""),
                                                          "bijgewerkt": str(r.get("Bijgewerkt") or "")}
              for r in lees("Uitleg") if r.get("Onderwerp")}

    inleg = [c for c in constructies.values() if any(st["categorie"].lower() == "inleg" for st in c["stromen"])]
    echte_constructies = [c for c in constructies.values() if c not in inleg]
    totaal_vast = round(sum(l["per_maand"] for l in lasten) + sum(p["per_maand"] for p in polissen)
                        + hyp["maandlast"] + sum(c["uit_pm"] for c in echte_constructies), 2)
    totaal_in = round(sum(c["in_pm"] for c in echte_constructies), 2)
    inkomen_pm = round(sum(i["per_maand"] for i in inkomsten), 2)
    sparen = [x for x in variabel if any(w in x["naam"].lower() for w in ("spaar", "sparen", "beleg", "pensioen"))]
    sparen_pm = round(sum(x["per_maand"] for x in sparen), 2)
    variabel_pm = round(sum(x["per_maand"] for x in variabel) - sparen_pm, 2)  # sparen is een keuze, geen kostenpost
    reservering_pm = round((inc_uit - inc_in) / 12, 2)  # wat je per maand opzij zou zetten voor voorspelbare pieken

    return {
        "beschikbaar": True, "link": link, "vandaag": vandaag.isoformat(),
        "vaste_lasten": sorted(lasten, key=lambda l: -l["per_maand"]),
        "per_categorie": dict(sorted(per_cat.items(), key=lambda kv: -kv[1])),
        "polissen": sorted(polissen, key=lambda p: (p["dagen"] is None, p["dagen"] or 0)),
        "hypotheek": hyp,
        "rekeningen": rekeningen,
        "inkomsten": sorted(inkomsten, key=lambda i: -i["per_maand"]),
        "variabel": sorted(variabel, key=lambda x: -x["per_maand"]),
        "incidenteel": sorted(incidenteel, key=lambda i: (i["dagen"] is None, i["dagen"] or 0)),
        "besparingen": besparingen,
        "via_inleg": via_inleg,
        "constructies": echte_constructies,
        "inleg": inleg,
        "verrekening": {"personen": personen, "saldo": saldo, "regels": regels, "tekst": verreken_tekst,
                        "verdeling": verdeling},
        "totalen": {"vast_pm": totaal_vast, "in_pm": totaal_in, "netto_pm": round(totaal_vast - totaal_in, 2),
                    "inkomen_pm": inkomen_pm, "variabel_pm": variabel_pm, "sparen_pm": sparen_pm,
                    "over_pm": round(inkomen_pm - (totaal_vast - totaal_in) - variabel_pm, 2),
                    "incidenteel_in_jaar": inc_in, "incidenteel_uit_jaar": inc_uit, "reservering_pm": reservering_pm,
                    "over_na_reservering_pm": round(inkomen_pm - (totaal_vast - totaal_in) - variabel_pm - reservering_pm, 2),
                    "lasten_pm": round(sum(l["per_maand"] for l in lasten), 2),
                    "polissen_pm": round(sum(p["per_maand"] for p in polissen), 2)},
        "uitleg": uitleg,
        "signalen": signalen(lasten, polissen, delen, vandaag),
        "woordenlijst": WOORDENLIJST,
    }


def signalen(lasten: list[dict], polissen: list[dict], delen: list[dict], vandaag: date) -> list[dict]:
    """Zonder bedragen (komt op het muurbord): opzegmomenten en einde rentevaste periode."""
    out = []
    for item, soort in [(l, "abonnement") for l in lasten] + [(p, "polis") for p in polissen]:
        if item["dagen"] is None:
            continue
        termijn = _termijn_dagen(item["opzegtermijn"])
        opzeg_voor = item["dagen"] - termijn  # dagen tot de laatste opzegdag
        if item["dagen"] < 0:
            out.append({"tekst": f"💶 {item['naam']} is verlopen ({-item['dagen']} dagen geleden) — verlengd of stopgezet?",
                        "l2": "geld", "ernst": 0})
        elif opzeg_voor <= 21:
            wanneer = "vandaag" if opzeg_voor <= 0 else f"over {opzeg_voor} dagen"
            out.append({"tekst": f"💶 {item['naam']} ({soort}): laatste dag om op te zeggen is {wanneer}"
                                 + (f" (opzegtermijn {item['opzegtermijn']})" if item["opzegtermijn"] else ""),
                        "l2": "geld", "ernst": 0 if opzeg_voor <= 7 else 1})
    for item in lasten + polissen:
        od = item.get("overstap_dagen")
        if od is not None and -14 <= od <= 30:
            out.append({"tekst": f"🔄 {item['naam']}: overstapmoment {'vandaag' if od == 0 else ('over ' + str(od) + ' dagen') if od > 0 else str(-od) + ' dagen geleden'} — vergelijken loont",
                        "l2": "geld", "ernst": 0 if od <= 7 else 1})
    for d in delen:
        if d["rentevast_dagen"] is not None and 0 <= d["rentevast_dagen"] <= 180:
            mnd = max(1, round(d["rentevast_dagen"] / 30))
            out.append({"tekst": f"🏠 Rentevaste periode van hypotheek {d['deel']} loopt over {mnd} maand{'en' if mnd > 1 else ''} af — tijd om rentes te vergelijken",
                        "l2": "geld", "ernst": 1 if d["rentevast_dagen"] > 60 else 0})
    return out


def toon() -> str:
    """Compacte tekst voor het brein (leg-uit-vragen, controle van het register)."""
    r = register()
    if not r["beschikbaar"]:
        return f"Geen register gevonden. Maak hem aan met: python /app/agent/financien.py maak (pad: {SHEET_PAD})"
    regels = [f"FINANCIEEL OVERZICHT (per maand, stand {r['vandaag']}) · Sheet: {r['link']}",
              f"Vaste lasten totaal € {r['totalen']['vast_pm']:,.2f}/mnd (lasten {r['totalen']['lasten_pm']:,.2f} · "
              f"polissen {r['totalen']['polissen_pm']:,.2f} · hypotheek {r['hypotheek']['maandlast']:,.2f}) · "
              f"terugkomend € {r['totalen']['in_pm']:,.2f}/mnd → netto € {r['totalen']['netto_pm']:,.2f}/mnd", ""]
    if r["inkomsten"]:
        regels.append(f"INKOMSTEN € {r['totalen']['inkomen_pm']:,.2f}/mnd → na vaste lasten en budgetten blijft over € {r['totalen']['over_pm']:,.2f}/mnd")
        for i in r["inkomsten"]:
            regels.append(f"• {i['naam']} — € {i['bedrag']:,.2f} per {i['frequentie']} (= {i['per_maand']:,.2f}/mnd) · komt binnen op {i['komt_binnen_op']} · hoort bij {i['hoort_bij']}")
    regels.append("VASTE LASTEN")
    for l in r["vaste_lasten"]:
        regels.append(f"• {l['naam']} — {l['categorie']} · € {l['bedrag']:,.2f} per {l['frequentie']} (= {l['per_maand']:,.2f}/mnd) · "
                      f"betaald van {l['betaald_van']} · hoort bij {l['hoort_bij']}"
                      + (f" · einddatum {l['einddatum']}" if l["einddatum"] else "")
                      + (f" · opzegtermijn {l['opzegtermijn']}" if l["opzegtermijn"] else ""))
    if r["variabel"]:
        regels.append(f"VARIABELE KOSTEN (budget) € {r['totalen']['variabel_pm']:,.2f}/mnd")
        for x in r["variabel"]:
            regels.append(f"• {x['naam']} — € {x['per_maand']:,.2f}/mnd · betaald van {x['betaald_van']} · hoort bij {x['hoort_bij']}")
    regels.append("POLISSEN")
    for p in r["polissen"]:
        regels.append(f"• {p['naam']} ({p['verzekeraar']}) — {p['dekking']} · premie € {p['premie']:,.2f} per {p['frequentie']} · "
                      f"eigen risico € {p['eigen_risico']:,.0f} · betaald van {p['betaald_van']} · hoort bij {p['hoort_bij']}"
                      + (f" · einddatum {p['einddatum']} (opzegtermijn {p['opzegtermijn'] or '?'})" if p["einddatum"] else ""))
    regels.append("HYPOTHEEK")
    for d in r["hypotheek"]["delen"]:
        regels.append(f"• {d['deel']} ({d['verstrekker']}) — {d['vorm']} · restschuld € {d['restschuld']:,.0f} van € {d['hoofdsom']:,.0f} · "
                      f"rente {d['rente_pct']}% vast tot {d['rentevast_tot'] or '?'} · maandlast € {d['maandlast']:,.2f} "
                      f"(rente {d['rente']:,.2f} + aflossing {d['aflossing']:,.2f})")
    regels.append("GELDSTROMEN / CONSTRUCTIES")
    for c in r["constructies"]:
        regels.append(f"• {c['naam']}: uit € {c['uit_pm']:,.2f}/mnd, in € {c['in_pm']:,.2f}/mnd → netto {c['netto_pm']:+,.2f}/mnd"
                      + (f" · {c['uitleg']}" if c["uitleg"] else ""))
    for c in r["inleg"]:
        regels.append(f"• INLEG {c['naam']}: € {c['uit_pm']:,.2f}/mnd (afspraak, telt niet mee in verrekenen)")
    for x in r["via_inleg"]:
        regels.append(f"  - zit in de inleg: {x['wat']} € {x['bedrag']:,.2f}/mnd · {x['tekst']}")
    if r["incidenteel"]:
        regels.append(f"INCIDENTEEL (per jaar): voorspelbaar in € {r['totalen']['incidenteel_in_jaar']:,.0f} · uit € {r['totalen']['incidenteel_uit_jaar']:,.0f} → reservering € {r['totalen']['reservering_pm']:,.2f}/mnd")
        for i in r["incidenteel"]:
            regels.append(f"• {i['naam']} — {i['soort']} € {i['bedrag']:,.2f} · {i['wanneer']} · {'voorspelbaar' if i['voorspelbaar'] else 'onvoorspelbaar'} · {i['rekening']} · hoort bij {i['hoort_bij']}")
    if r["besparingen"]:
        regels.append("BESPARINGSVOORSTELLEN")
        for b in r["besparingen"]:
            regels.append(f"• {b['voorstel']} — € {b['per_maand']:,.2f}/mnd · {b['status']} · {b['categorie']}" + (f" · {b['notitie']}" if b["notitie"] else ""))
    v = r["verrekening"]
    regels.append(f"VERREKENING (structureel, per maand, t.o.v. de pot): {v['tekst']}")
    for rg in v["regels"]:
        regels.append(f"  - {rg['wat']}: € {rg['bedrag']:,.2f} · {rg['tekst']}")
    for s in r["signalen"]:
        regels.append(f"SIGNAAL: {s['tekst']}")
    return "\n".join(regels)


def main() -> None:
    cmd = sys.argv[1] if len(sys.argv) > 1 else "toon"
    if cmd == "maak":
        print("Sheet:", maak_sheet())
    elif cmd == "toon":
        print(toon())
    else:
        sys.exit("gebruik: financien.py toon | maak")


if __name__ == "__main__":
    if __package__ is None or __package__ == "":  # los script: python /app/agent/financien.py
        sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
        from agent import financien as _zelf  # noqa: F401
        sys.exit(_zelf.main())
    main()
